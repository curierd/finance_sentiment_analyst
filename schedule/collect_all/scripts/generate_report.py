#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generate quantitative Markdown report from the database.

Reads comments where created_at falls within the window:
  2026-06-18 15:00:00 ~ 2026-06-19 09:30:00 CST

Uses sentiment data already populated by db/update_sentiment.py.

Outputs:
  - schedule/collect_all/output/sentiment-report-<date>.md
"""
import argparse
import io
import json
import sys
from datetime import datetime, timezone, timedelta
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

SCRIPT_DIR = Path(__file__).resolve().parent
SCHEDULE_DIR = SCRIPT_DIR.parent
REPO_ROOT = SCHEDULE_DIR.parent.parent
sys.path.insert(0, str(REPO_ROOT))

from backend.database import get_db  # noqa: E402

CST = timezone(timedelta(hours=8))


def log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}", flush=True)


def fetch_in_window(window_start, window_end):
    """Fetch all comments whose created_at falls in [start, end] CST.
    created_at stored as ISO 8601 UTC with 'Z' suffix or naive ISO.
    """
    ws_utc = window_start.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    we_utc = window_end.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    conn = get_db()
    rows = conn.execute(
        """
        SELECT id, platform, comment_id, author_name, content,
               likes, replies, retweets, source_url, symbol,
               created_at, sentiment, sentiment_score,
               COALESCE(sentiment_fix, sentiment) AS effective_sentiment
        FROM comments
        WHERE created_at IS NOT NULL
          AND datetime(created_at) >= datetime(?, '-8 hours')
          AND datetime(created_at) <= datetime(?, '-8 hours')
        ORDER BY platform, datetime(created_at)
        """,
        (ws_utc, we_utc),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def summarize(records):
    """Return summary statistics per platform + overall."""
    out = {"total": len(records), "by_platform": {}, "overall": {}}
    counts = {"正面": 0, "中性": 0, "负面": 0}
    score_sum = 0.0
    high_like_total = 0
    high_like_pos = 0
    high_like_neg = 0
    for r in records:
        s = r.get("effective_sentiment") or "中性"
        counts[s] = counts.get(s, 0) + 1
        score = r.get("sentiment_score") or 0.0
        try:
            score = float(score)
        except (TypeError, ValueError):
            score = 0.0
        score_sum += score
        likes = r.get("likes") or 0
        if likes > 10:
            high_like_total += 1
            if s == "正面":
                high_like_pos += 1
            elif s == "负面":
                high_like_neg += 1
    if out["total"] > 0:
        out["overall"] = {
            "positive": counts.get("正面", 0),
            "neutral": counts.get("中性", 0),
            "negative": counts.get("负面", 0),
            "positive_pct": round(100 * counts.get("正面", 0) / out["total"], 1),
            "neutral_pct": round(100 * counts.get("中性", 0) / out["total"], 1),
            "negative_pct": round(100 * counts.get("负面", 0) / out["total"], 1),
            "score_avg": round(score_sum / out["total"], 3),
            "score_sum": round(score_sum, 3),
            "high_like_total": high_like_total,
            "high_like_pos": high_like_pos,
            "high_like_neg": high_like_neg,
        }

    # Per-platform
    by_p = {}
    for r in records:
        p = r.get("platform", "unknown")
        by_p.setdefault(p, []).append(r)
    for p, lst in by_p.items():
        c = {"正面": 0, "中性": 0, "负面": 0}
        ss = 0.0
        hl_total = 0
        hl_pos = 0
        hl_neg = 0
        for r in lst:
            s = r.get("effective_sentiment") or "中性"
            c[s] = c.get(s, 0) + 1
            score = r.get("sentiment_score") or 0.0
            try:
                score = float(score)
            except (TypeError, ValueError):
                score = 0.0
            ss += score
            likes = r.get("likes") or 0
            if likes > 10:
                hl_total += 1
                if s == "正面":
                    hl_pos += 1
                elif s == "负面":
                    hl_neg += 1
        n = len(lst)
        by_p[p] = {
            "total": n,
            "positive": c.get("正面", 0),
            "neutral": c.get("中性", 0),
            "negative": c.get("负面", 0),
            "positive_pct": round(100 * c.get("正面", 0) / n, 1) if n else 0,
            "neutral_pct": round(100 * c.get("中性", 0) / n, 1) if n else 0,
            "negative_pct": round(100 * c.get("负面", 0) / n, 1) if n else 0,
            "score_avg": round(ss / n, 3) if n else 0.0,
            "high_like_total": hl_total,
            "high_like_pos": hl_pos,
            "high_like_neg": hl_neg,
        }
    out["by_platform"] = by_p
    return out


def render_markdown(summary, window_start, window_end, today):
    lines = []
    lines.append(f"# 全平台散户情绪报告 — {today}")
    lines.append("")
    lines.append(f"- 时间窗口（CST）：{window_start.isoformat()} ~ {window_end.isoformat()}")
    lines.append(f"- 总评论数：**{summary['total']}**")
    lines.append("")
    lines.append("## 全平台汇总")
    o = summary["overall"]
    lines.append(f"| 维度 | 正面 | 中性 | 负面 | 平均得分 |")
    lines.append(f"|------|------|------|------|----------|")
    lines.append(
        f"| 全部评论 | {o['positive']} ({o['positive_pct']}%) | "
        f"{o['neutral']} ({o['neutral_pct']}%) | "
        f"{o['negative']} ({o['negative_pct']}%) | "
        f"{o['score_avg']} |"
    )
    lines.append(
        f"| 高赞 (likes>10) | {o['high_like_pos']} | 0 | "
        f"{o['high_like_neg']} | — |"
    )
    lines.append("")

    lines.append("## 按平台拆分")
    lines.append("")
    lines.append("| 平台 | 评论数 | 正面% | 中性% | 负面% | 平均得分 | 高赞 | 高赞正面 | 高赞负面 |")
    lines.append("|------|--------|-------|-------|-------|----------|------|----------|----------|")
    for p, s in summary["by_platform"].items():
        lines.append(
            f"| {p} | {s['total']} | {s['positive_pct']}% | "
            f"{s['neutral_pct']}% | {s['negative_pct']}% | "
            f"{s['score_avg']} | {s['high_like_total']} | "
            f"{s['high_like_pos']} | {s['high_like_neg']} |"
        )
    lines.append("")

    # Interpretation
    lines.append("## 解读")
    lines.append("")
    score_avg = o["score_avg"]
    pos_pct = o["positive_pct"]
    neg_pct = o["negative_pct"]
    if score_avg > 0.05 and pos_pct > neg_pct:
        mood = "偏多"
    elif score_avg < -0.05 and neg_pct > pos_pct:
        mood = "偏空"
    else:
        mood = "中性观望"
    lines.append(
        f"- 综合得分 {score_avg}（正值偏多 / 负值偏空），整体情绪：**{mood}**"
    )
    lines.append(
        f"- 正面占比 {pos_pct}% vs 负面占比 {neg_pct}%，差值 {round(pos_pct - neg_pct, 1)} 个百分点"
    )
    if o["high_like_total"]:
        hl_ratio = round(100 * o["high_like_pos"] / o["high_like_total"], 1)
        lines.append(
            f"- 高赞评论 {o['high_like_total']} 条，正面 {o['high_like_pos']} 条（{hl_ratio}%），"
            f"被广泛认同的声音同样{mood}"
        )
    lines.append("")
    lines.append("## 数据源")
    lines.append("")
    lines.append("- 评论表：`db/comments.db.comments`（按 `created_at` 在窗口内过滤）")
    lines.append("- 情绪模型：`jobs/sentiment_analyzer/llm_sentiment.SentimentAnalyzer`（LLM / DeepSeek V3）")
    lines.append("- 入库脚本：")
    lines.append("  - `jobs/bilibili_comments_collector/scripts/collect_bilibili_today.py`")
    lines.append("  - `jobs/xiaohongshu_comments_collector/scripts/import_to_db.py`")
    lines.append("  - `schedule/collect_all/scripts/import_xueqiu_to_db.py`")
    lines.append("")
    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", required=True)
    parser.add_argument("--window-start", required=True)
    parser.add_argument("--window-end", required=True)
    args = parser.parse_args()

    window_start = datetime.fromisoformat(args.window_start)
    window_end = datetime.fromisoformat(args.window_end)

    log(f"Generating report for window {window_start} ~ {window_end}")

    records = fetch_in_window(window_start, window_end)
    log(f"Fetched {len(records)} comments in window")
    summary = summarize(records)

    out_dir = SCHEDULE_DIR / "output"
    out_dir.mkdir(exist_ok=True)
    out_file = out_dir / f"sentiment-report-{args.date}.md"
    md = render_markdown(summary, window_start, window_end, args.date)
    out_file.write_text(md, encoding="utf-8")

    # Also dump summary JSON for downstream tooling
    summary_file = out_dir / f"sentiment-summary-{args.date}.json"
    summary_file.write_text(
        json.dumps(
            {
                "window_start": window_start.isoformat(),
                "window_end": window_end.isoformat(),
                "summary": summary,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    log(f"Saved: {out_file}")
    log(f"Saved: {summary_file}")

    # Excel
    xlsx_file = out_dir / f"sentiment-report-{args.date}.xlsx"
    write_excel(xlsx_file, summary, window_start, window_end, args.date, records)
    log(f"Saved: {xlsx_file}")

    # HTML
    html_file = out_dir / f"sentiment-report-{args.date}.html"
    html = render_html(summary, window_start, window_end, args.date, records)
    html_file.write_text(html, encoding="utf-8")
    log(f"Saved: {html_file}")


def write_excel(path, summary, window_start, window_end, today, records):
    """Write a multi-sheet Excel workbook.

    Sheets:
      - Summary: overall + per-platform stats
      - Comments: full per-comment records within the window
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    pos_fill = PatternFill("solid", fgColor="C6EFCE")
    neg_fill = PatternFill("solid", fgColor="FFC7CE")
    neu_fill = PatternFill("solid", fgColor="FFEB9C")
    thin = Side(border_style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title block
    ws["A1"] = f"全平台散户情绪报告 — {today}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:H1")
    ws["A2"] = "时间窗口（CST）"
    ws["B2"] = f"{window_start.isoformat()} ~ {window_end.isoformat()}"
    ws["A3"] = "总评论数"
    ws["B3"] = summary["total"]

    # Overall summary
    ws["A5"] = "全平台汇总"
    ws["A5"].font = Font(bold=True, size=12)
    headers = ["维度", "正面", "正面%", "中性", "中性%", "负面", "负面%", "平均得分"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=6, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = border
    o = summary["overall"]
    row = 7
    ws.cell(row=row, column=1, value="全部评论").border = border
    for col, v in enumerate(
        [o["positive"], o["positive_pct"], o["neutral"], o["neutral_pct"],
         o["negative"], o["negative_pct"], o["score_avg"]],
        start=2,
    ):
        c = ws.cell(row=row, column=col, value=v)
        c.border = border
        if col == 2:
            c.fill = pos_fill
        elif col == 4:
            c.fill = neu_fill
        elif col == 6:
            c.fill = neg_fill

    row = 8
    ws.cell(row=row, column=1, value="高赞(likes>10)").border = border
    for col, v in enumerate(
        [o["high_like_pos"], None, None, None, o["high_like_neg"], None, None],
        start=2,
    ):
        c = ws.cell(row=row, column=col, value=v)
        c.border = border
        if col == 2:
            c.fill = pos_fill
        elif col == 6:
            c.fill = neg_fill

    # Per-platform
    row = 10
    ws.cell(row=row, column=1, value="按平台拆分").font = Font(bold=True, size=12)
    row = 11
    plat_headers = ["平台", "评论数", "正面", "正面%", "中性", "中性%",
                    "负面", "负面%", "平均得分", "高赞", "高赞正面", "高赞负面"]
    for col, h in enumerate(plat_headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = border

    for p, s in summary["by_platform"].items():
        row += 1
        vals = [p, s["total"], s["positive"], s["positive_pct"],
                s["neutral"], s["neutral_pct"],
                s["negative"], s["negative_pct"],
                s["score_avg"], s["high_like_total"],
                s["high_like_pos"], s["high_like_neg"]]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=row, column=col, value=v)
            c.border = border
            if col == 4:
                c.fill = pos_fill
            elif col == 6:
                c.fill = neu_fill
            elif col == 8:
                c.fill = neg_fill

    # Auto width
    for col_idx in range(1, 13):
        letter = ws.cell(row=11, column=col_idx).column_letter
        ws.column_dimensions[letter].width = 14
    ws.column_dimensions["A"].width = 22

    # Comments sheet
    ws2 = wb.create_sheet("Comments")
    comments_headers = ["platform", "author_name", "created_at", "likes",
                        "replies", "sentiment", "sentiment_score", "content"]
    for col, h in enumerate(comments_headers, start=1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = border

    row = 1
    for r in records:
        row += 1
        sentiment = r.get("effective_sentiment") or "中性"
        try:
            score = float(r.get("sentiment_score") or 0.0)
        except (TypeError, ValueError):
            score = 0.0
        vals = [
            r.get("platform", ""),
            r.get("author_name", ""),
            r.get("created_at", ""),
            r.get("likes", 0) or 0,
            r.get("replies", 0) or 0,
            sentiment,
            score,
            (r.get("content", "") or "").replace("\n", " "),
        ]
        for col, v in enumerate(vals, start=1):
            c = ws2.cell(row=row, column=col, value=v)
            c.border = border
            if col == 6:
                if sentiment == "正面":
                    c.fill = pos_fill
                elif sentiment == "负面":
                    c.fill = neg_fill
                else:
                    c.fill = neu_fill
    # Widths
    widths = [12, 18, 22, 8, 8, 10, 14, 80]
    for i, w in enumerate(widths, start=1):
        ws2.column_dimensions[ws2.cell(row=1, column=i).column_letter].width = w
    ws2.freeze_panes = "A2"

    wb.save(path)


def render_html(summary, window_start, window_end, today, records):
    """Render a self-contained HTML report with embedded styles."""
    o = summary["overall"]
    score_avg = o["score_avg"]
    pos_pct = o["positive_pct"]
    neg_pct = o["negative_pct"]
    if score_avg > 0.05 and pos_pct > neg_pct:
        mood = "偏多"
        mood_color = "#10b981"
    elif score_avg < -0.05 and neg_pct > pos_pct:
        mood = "偏空"
        mood_color = "#ef4444"
    else:
        mood = "中性观望"
        mood_color = "#6b7280"

    rows = ""
    for r in records[:500]:  # cap at 500 in HTML for readability
        sentiment = r.get("effective_sentiment") or "中性"
        color = {"正面": "#10b981", "中性": "#6b7280", "负面": "#ef4444"}.get(sentiment, "#6b7280")
        content = (r.get("content", "") or "").replace("<", "&lt;").replace(">", "&gt;")
        rows += (
            f"<tr>"
            f"<td>{r.get('platform','')}</td>"
            f"<td>{(r.get('author_name','') or '')[:20]}</td>"
            f"<td>{(r.get('created_at','') or '')[:19]}</td>"
            f"<td style='text-align:right'>{r.get('likes', 0) or 0}</td>"
            f"<td style='color:{color};font-weight:600'>{sentiment}</td>"
            f"<td class='content'>{content[:200]}</td>"
            f"</tr>"
        )

    plat_rows = ""
    for p, s in summary["by_platform"].items():
        plat_rows += (
            f"<tr><td>{p}</td><td>{s['total']}</td>"
            f"<td>{s['positive']} ({s['positive_pct']}%)</td>"
            f"<td>{s['neutral']} ({s['neutral_pct']}%)</td>"
            f"<td>{s['negative']} ({s['negative_pct']}%)</td>"
            f"<td>{s['score_avg']}</td>"
            f"<td>{s['high_like_total']}</td>"
            f"<td>{s['high_like_pos']}</td>"
            f"<td>{s['high_like_neg']}</td></tr>"
        )

    high_like_ratio = (
        round(100 * o["high_like_pos"] / o["high_like_total"], 1)
        if o["high_like_total"] else 0
    )

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<title>全平台散户情绪报告 — {today}</title>
<style>
  body {{
    font-family: -apple-system, "Segoe UI", "Microsoft YaHei", sans-serif;
    margin: 0; padding: 24px; background: #f9fafb; color: #111827;
  }}
  .container {{ max-width: 1200px; margin: 0 auto; background: #fff;
                border-radius: 12px; padding: 32px; box-shadow: 0 1px 3px rgba(0,0,0,.06); }}
  h1 {{ margin: 0 0 8px; font-size: 24px; }}
  h2 {{ margin-top: 32px; font-size: 18px; border-bottom: 2px solid #e5e7eb; padding-bottom: 8px; }}
  .meta {{ color: #6b7280; font-size: 14px; margin-bottom: 24px; }}
  .kpis {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 16px; margin: 16px 0 32px; }}
  .kpi {{ background: #f3f4f6; border-radius: 8px; padding: 16px; }}
  .kpi .label {{ color: #6b7280; font-size: 12px; }}
  .kpi .value {{ font-size: 24px; font-weight: 700; margin-top: 4px; }}
  .kpi.positive .value {{ color: #10b981; }}
  .kpi.neutral .value {{ color: #6b7280; }}
  .kpi.negative .value {{ color: #ef4444; }}
  .kpi.mood .value {{ color: {mood_color}; }}
  table {{ width: 100%; border-collapse: collapse; margin-top: 12px; font-size: 14px; }}
  th, td {{ padding: 8px 12px; border-bottom: 1px solid #e5e7eb; text-align: left; }}
  th {{ background: #1f4e78; color: #fff; }}
  td.content {{ max-width: 480px; overflow: hidden; text-overflow: ellipsis; }}
  tr:hover td {{ background: #f9fafb; }}
  .mood-pill {{ display: inline-block; padding: 4px 16px; border-radius: 9999px;
                background: {mood_color}; color: #fff; font-weight: 600; }}
  .note {{ color: #6b7280; font-size: 12px; margin-top: 16px; }}
  .legend {{ color: #6b7280; font-size: 12px; }}
</style>
</head>
<body>
<div class="container">
  <h1>全平台散户情绪报告 — {today}</h1>
  <div class="meta">
    时间窗口（CST）：{window_start.isoformat()} ~ {window_end.isoformat()}
    &nbsp;·&nbsp; 报告生成：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
  </div>

  <div class="kpis">
    <div class="kpi"><div class="label">总评论数</div><div class="value">{summary['total']}</div></div>
    <div class="kpi positive"><div class="label">正面</div><div class="value">{o['positive']} ({pos_pct}%)</div></div>
    <div class="kpi neutral"><div class="label">中性</div><div class="value">{o['neutral']} ({o['neutral_pct']}%)</div></div>
    <div class="kpi negative"><div class="label">负面</div><div class="value">{o['negative']} ({neg_pct}%)</div></div>
  </div>

  <h2>整体情绪 <span class="mood-pill">{mood}</span></h2>
  <ul>
    <li>综合得分 <strong>{score_avg}</strong>（正值偏多 / 负值偏空）</li>
    <li>正面 {pos_pct}% vs 负面 {neg_pct}%，差值 {round(pos_pct - neg_pct, 1)} 个百分点</li>
    <li>高赞评论 (likes&gt;10) {o['high_like_total']} 条，正面 {o['high_like_pos']} 条（{high_like_ratio}%），负面 {o['high_like_neg']} 条</li>
  </ul>

  <h2>按平台拆分</h2>
  <table>
    <thead>
      <tr>
        <th>平台</th><th>评论数</th><th>正面</th><th>中性</th><th>负面</th>
        <th>平均得分</th><th>高赞</th><th>高赞正面</th><th>高赞负面</th>
      </tr>
    </thead>
    <tbody>
      {plat_rows}
    </tbody>
  </table>

  <h2>评论明细 <span class="legend">(最多展示 500 条)</span></h2>
  <table>
    <thead>
      <tr>
        <th>平台</th><th>作者</th><th>时间</th><th>点赞</th><th>情绪</th><th>内容</th>
      </tr>
    </thead>
    <tbody>
      {rows}
    </tbody>
  </table>

  <p class="note">
    数据源：<code>db/comments.db.comments</code>（按 <code>created_at</code> 在窗口内过滤）。
    情绪模型：<code>jobs/sentiment_analyzer/llm_sentiment.SentimentAnalyzer</code>（LLM / DeepSeek V3）。
  </p>
</div>
</body>
</html>
"""
    return html


if __name__ == "__main__":
    main()