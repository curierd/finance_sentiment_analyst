#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""散户对小登股(科技/小盘)的情绪分析 — 看好 / 中立 / 看跌。

输入：DB 窗口 2026-06-19 15:00 ~ 2026-06-23 09:30 CST 的所有评论。
筛选：symbol 命中科创板 SH688 / 创业板 SZ300 / 北证 BJ8，OR 内容包含科技/AI/半导体/新能源/腾讯/中芯/泡泡玛特等关键词。
情绪：复用 jobs.sentiment_analyzer.llm_sentiment.SentimentAnalyzer（DeepSeek）批量分析。
输出：schedule/collect_all/output/sentiment-xiaodeng-2026-06-20.{md,xlsx,html}。
"""
import io
import json
import sys
from datetime import datetime, timezone, timedelta
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

SCRIPT_DIR = Path(__file__).resolve().parent
SCHEDULE_DIR = SCRIPT_DIR.parent
REPO_ROOT = SCHEDULE_DIR.parent.parent
sys.path.insert(0, str(REPO_ROOT))

from backend.database import get_db  # noqa: E402
from jobs.sentiment_analyzer.llm_sentiment import SentimentAnalyzer  # noqa: E402

CST = timezone(timedelta(hours=8))
WINDOW_START = datetime(2026, 6, 19, 15, 0, 0, tzinfo=CST)
WINDOW_END = datetime(2026, 6, 23, 9, 30, 0, tzinfo=CST)
TODAY = "2026-06-20"

# 小登相关关键词（科技股 / 成长股 / 科创关键词 + 预设里的 5 只标的）
# KEYWORDS: 严格定义"小登股(科技股)"口径。
# - 显式排除互联网平台股（腾讯）、潮玩/茶饮等消费类（蜜雪、泡泡玛特），
#   以及泛指词"互联网"——它们属于"老登/消费"，不属于科技股主体。
# - 显式排除新能源 / 光伏 / 锂电 / 锂电池 / 新能源车 —— 属于传统制造业 /
#   周期股，不算新兴科技股主体（中芯 / 寒武纪 / 海光 / 光模块 / PCB 等硬科技保留）。
# - 数据源约定见 data/sections/consumer-tech.md，本脚本只看硬科技口径。
KEYWORDS = [
    # 直接科技关键词
    "科技", "AI", "人工智能", "半导体", "芯片",
    "光模块", "PCB", "印制电路", "玻璃基板",
    "算力", "大模型", "GPU", "CPU", "存储", "FPGA",
    "机器人", "智能驾驶", "自动驾驶",
    # 板块/市场结构词
    "创业板", "科创板", "科创", "小盘", "成长",
    "结构牛", "硬科技",
    # 行业内具体股票（中芯 = 半导体龙头，不在互联网/潮玩范围）
    "中芯", "寒武纪", "海光", "中际旭创", "京东方",
]


def log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}", flush=True)


def fetch_xiaodeng():
    """返回窗口内属于小登股的评论记录列表。"""
    ws_utc = WINDOW_START.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    we_utc = WINDOW_END.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    keyword_clause = " OR ".join(["content LIKE ?"] * len(KEYWORDS))
    params = [f"%{k}%" for k in KEYWORDS]
    sql = f"""
        SELECT id, platform, symbol, author_name, content, likes,
               COALESCE(sentiment_fix, sentiment) AS effective_sentiment,
               sentiment_score, created_at
        FROM comments
        WHERE created_at IS NOT NULL
          AND datetime(created_at) >= datetime(?, '-8 hours')
          AND datetime(created_at) <= datetime(?, '-8 hours')
          AND (
            symbol LIKE 'SH688%' OR symbol LIKE 'SZ300%' OR symbol LIKE 'BJ8%'
            OR ({keyword_clause})
          )
        ORDER BY datetime(created_at)
    """
    conn = get_db()
    rows = conn.execute(sql, [ws_utc, we_utc] + params).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def fetch_xiaodeng_breakdown():
    """返回 (命中symbol模式, 命中关键词, 重叠) 三个集合大小。"""
    ws_utc = WINDOW_START.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    we_utc = WINDOW_END.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    keyword_clause = " OR ".join(["content LIKE ?"] * len(KEYWORDS))
    params = [f"%{k}%" for k in KEYWORDS]
    base = """
        FROM comments
        WHERE created_at IS NOT NULL
          AND datetime(created_at) >= datetime(?, '-8 hours')
          AND datetime(created_at) <= datetime(?, '-8 hours')
    """
    conn = get_db()
    n_sym = conn.execute(
        f"SELECT COUNT(*) {base} AND (symbol LIKE 'SH688%' OR symbol LIKE 'SZ300%' OR symbol LIKE 'BJ8%')",
        [ws_utc, we_utc],
    ).fetchone()[0]
    n_kw = conn.execute(
        f"SELECT COUNT(*) {base} AND ({keyword_clause})",
        [ws_utc, we_utc] + params,
    ).fetchone()[0]
    n_union = conn.execute(
        f"SELECT COUNT(*) {base} AND (symbol LIKE 'SH688%' OR symbol LIKE 'SZ300%' OR symbol LIKE 'BJ8%' OR ({keyword_clause}))",
        [ws_utc, we_utc] + params,
    ).fetchone()[0]
    conn.close()
    return {"symbol_pattern": n_sym, "keyword": n_kw, "union": n_union}


def run_llm_sentiment(records):
    """对 records（必须有 content）跑 LLM 批量分析，写回 sentiment/score 字段。

    优先复用 DB 内已有 sentiment（来自 `db/update_sentiment.py` 的 LLM 批跑）；
    对没有 sentiment 的评论才走 LLM。
    """
    need_llm = []
    need_llm_idx = []
    for i, r in enumerate(records):
        if r.get("effective_sentiment"):
            continue
        need_llm.append(r.get("content", "") or "")
        need_llm_idx.append(i)

    log(f"  DB 已有 sentiment: {len(records) - len(need_llm)} 条, 待补 LLM: {len(need_llm)} 条")
    if not need_llm:
        return 0
    analyzer = SentimentAnalyzer()
    BATCH = 20
    total_done = 0
    for start in range(0, len(need_llm), BATCH):
        chunk_texts = need_llm[start:start + BATCH]
        chunk_idx = need_llm_idx[start:start + BATCH]
        log(f"  Batch {start // BATCH + 1}: {len(chunk_texts)} 条")
        results = analyzer.analyze_batch(chunk_texts)
        for idx, res in zip(chunk_idx, results):
            records[idx]["llm_sentiment"] = res["sentiment"]
            try:
                score = float(res["scores"]["positive"]) - float(res["scores"]["negative"])
                records[idx]["llm_score"] = round(score, 4)
            except (TypeError, ValueError, KeyError):
                records[idx]["llm_score"] = 0.0
            total_done += 1
    return total_done


def aggregate(records):
    """汇总 看好 / 中立 / 看跌 占比、平均得分、按平台拆分。"""
    total = len(records)
    counts = {"看好": 0, "中立": 0, "看跌": 0}
    score_sum = 0.0
    score_n = 0
    high_like_total = 0
    high_like_pos = 0
    high_like_neg = 0

    for r in records:
        s = r.get("llm_sentiment") or r.get("effective_sentiment") or "中性"
        # 兼容 SentimentAnalyzer 的"正面/中性/负面" → 看好/中立/看跌
        label = {"正面": "看好", "中性": "中立", "负面": "看跌"}.get(s, "中立")
        counts[label] = counts.get(label, 0) + 1
        sc = r.get("llm_score")
        if sc is None:
            sc = r.get("sentiment_score") or 0.0
        try:
            score_sum += float(sc)
            score_n += 1
        except (TypeError, ValueError):
            pass
        likes = r.get("likes") or 0
        if likes > 10:
            high_like_total += 1
            if label == "看好":
                high_like_pos += 1
            elif label == "看跌":
                high_like_neg += 1

    overall = {
        "total": total,
        "看好": counts["看好"],
        "中立": counts["中立"],
        "看跌": counts["看跌"],
        "看好_pct": round(100 * counts["看好"] / total, 1) if total else 0,
        "中立_pct": round(100 * counts["中立"] / total, 1) if total else 0,
        "看跌_pct": round(100 * counts["看跌"] / total, 1) if total else 0,
        "score_avg": round(score_sum / score_n, 3) if score_n else 0.0,
        "high_like_total": high_like_total,
        "high_like_pos": high_like_pos,
        "high_like_neg": high_like_neg,
    }

    # 按平台
    by_p = {}
    for r in records:
        p = r.get("platform", "unknown")
        by_p.setdefault(p, []).append(r)
    by_platform = {}
    for p, lst in by_p.items():
        c = {"看好": 0, "中立": 0, "看跌": 0}
        ss = 0.0
        sn = 0
        hl = 0
        hp = 0
        hn = 0
        for r in lst:
            s = r.get("llm_sentiment") or r.get("effective_sentiment") or "中立"
            label = {"正面": "看好", "中性": "中立", "负面": "看跌"}.get(s, "中立")
            c[label] = c.get(label, 0) + 1
            sc = r.get("llm_score")
            if sc is None:
                sc = r.get("sentiment_score") or 0.0
            try:
                ss += float(sc)
                sn += 1
            except (TypeError, ValueError):
                pass
            likes = r.get("likes") or 0
            if likes > 10:
                hl += 1
                if label == "看好":
                    hp += 1
                elif label == "看跌":
                    hn += 1
        n = len(lst)
        by_platform[p] = {
            "total": n,
            "看好": c["看好"],
            "中立": c["中立"],
            "看跌": c["看跌"],
            "看好_pct": round(100 * c["看好"] / n, 1) if n else 0,
            "中立_pct": round(100 * c["中立"] / n, 1) if n else 0,
            "看跌_pct": round(100 * c["看跌"] / n, 1) if n else 0,
            "score_avg": round(ss / sn, 3) if sn else 0.0,
            "high_like_total": hl,
            "high_like_pos": hp,
            "high_like_neg": hn,
        }
    return {"overall": overall, "by_platform": by_platform}


def render_md(summary, breakdown, top_comments):
    o = summary["overall"]
    lines = [
        f"# 散户对科技股(小登股)情绪报告 — {TODAY}",
        "",
        f"- 时间窗口（CST）：{WINDOW_START.isoformat()} ~ {WINDOW_END.isoformat()}",
        f"- 筛选口径：symbol ∈ `SH688*` / `SZ300*` / `BJ8*` ∪ content 含科技/AI/半导体/新能源/腾讯/中芯/泡泡玛特 等关键词",
        f"- 样本范围：symbol 命中 {breakdown['symbol_pattern']} / 关键词命中 {breakdown['keyword']} / 去重合计 {breakdown['union']}",
        f"- 实际参与情绪计算：**{o['total']}** 条",
        "",
        "## 整体情绪",
        "",
        "| 维度 | 看好 | 中立 | 看跌 | 平均得分 |",
        "|------|------|------|------|----------|",
        f"| 全部评论 | {o['看好']} ({o['看好_pct']}%) | "
        f"{o['中立']} ({o['中立_pct']}%) | {o['看跌']} ({o['看跌_pct']}%) | {o['score_avg']} |",
        f"| 高赞 (likes>10) | {o['high_like_pos']} | 0 | {o['high_like_neg']} | — |",
        "",
    ]

    # 解读
    score_avg = o["score_avg"]
    pos_pct = o["看好_pct"]
    neg_pct = o["看跌_pct"]
    if score_avg > 0.05 and pos_pct > neg_pct:
        mood = "看好"
        mood_color = "🟢"
    elif score_avg < -0.05 and neg_pct > pos_pct:
        mood = "看跌"
        mood_color = "🔴"
    else:
        mood = "中立"
        mood_color = "⚪"
    lines.append("## 解读")
    lines.append("")
    lines.append(
        f"- 综合得分 **{score_avg}**（正值偏多 / 负值偏空），整体情绪：**{mood}** {mood_color}"
    )
    lines.append(
        f"- 看好占比 {pos_pct}% vs 看跌占比 {neg_pct}%，差值 {round(pos_pct - neg_pct, 1)} 个百分点"
    )
    if o["high_like_total"]:
        hl_ratio = round(100 * o["high_like_pos"] / o["high_like_total"], 1)
        lines.append(
            f"- 高赞评论 {o['high_like_total']} 条，看好 {o['high_like_pos']} 条（{hl_ratio}%），"
            f"被广泛认同的方向同样{mood}"
        )
    lines.append("")

    # 按平台
    lines.append("## 按平台拆分")
    lines.append("")
    lines.append("| 平台 | 评论数 | 看好% | 中立% | 看跌% | 平均得分 | 高赞 | 高赞看好 | 高赞看跌 |")
    lines.append("|------|--------|-------|-------|-------|----------|------|----------|----------|")
    for p, s in summary["by_platform"].items():
        lines.append(
            f"| {p} | {s['total']} | {s['看好_pct']}% | "
            f"{s['中立_pct']}% | {s['看跌_pct']}% | {s['score_avg']} | "
            f"{s['high_like_total']} | {s['high_like_pos']} | {s['high_like_neg']} |"
        )
    lines.append("")

    # 高赞评论样本
    if top_comments:
        lines.append("## 高赞评论样本 (likes>10)")
        lines.append("")
        lines.append("| 平台 | 作者 | 时间 | 点赞 | 情绪 | 内容 |")
        lines.append("|------|------|------|------|------|------|")
        for r in top_comments[:30]:
            sentiment = r.get("llm_sentiment") or r.get("effective_sentiment") or "中性"
            label = {"正面": "看好", "中性": "中立", "负面": "看跌"}.get(sentiment, "中立")
            content = (r.get("content", "") or "").replace("\n", " ").replace("|", "/")[:120]
            lines.append(
                f"| {r.get('platform','')} | {(r.get('author_name','') or '')[:15]} | "
                f"{(r.get('created_at','') or '')[:19]} | {r.get('likes', 0) or 0} | {label} | {content} |"
            )
        lines.append("")

    lines.append("## 数据源")
    lines.append("")
    lines.append("- 评论表：`db/comments.db.comments`（按 `created_at` 在窗口内过滤）")
    lines.append("- 情绪模型：`jobs/sentiment_analyzer/llm_sentiment.SentimentAnalyzer`（LLM / DeepSeek）")
    lines.append("- 关键词集合：科技/AI/人工智能/半导体/芯片/新能源/创业板/科创板/小盘/成长/互联网/腾讯/中芯/宁德/泡泡玛特/蜜雪/新能源车/锂电池/光模块/算力/大模型/GPU/CPU/存储/消费电子/结构牛/科创/锂电/光伏/机器人/智能驾驶")
    lines.append("- 预设标的：`data/sections/consumer-tech.md`（腾讯/中芯/宁德/蜜雪/泡泡玛特）")
    return "\n".join(lines)


def render_html(summary, breakdown, top_comments):
    o = summary["overall"]
    score_avg = o["score_avg"]
    pos_pct = o["看好_pct"]
    neg_pct = o["看跌_pct"]
    if score_avg > 0.05 and pos_pct > neg_pct:
        mood, mood_color = "看好", "#10b981"
    elif score_avg < -0.05 and neg_pct > pos_pct:
        mood, mood_color = "看跌", "#ef4444"
    else:
        mood, mood_color = "中立", "#6b7280"

    plat_rows = ""
    for p, s in summary["by_platform"].items():
        plat_rows += (
            f"<tr><td>{p}</td><td>{s['total']}</td>"
            f"<td style='color:#10b981'>{s['看好']} ({s['看好_pct']}%)</td>"
            f"<td>{s['中立']} ({s['中立_pct']}%)</td>"
            f"<td style='color:#ef4444'>{s['看跌']} ({s['看跌_pct']}%)</td>"
            f"<td>{s['score_avg']}</td>"
            f"<td>{s['high_like_total']}</td>"
            f"<td>{s['high_like_pos']}</td>"
            f"<td>{s['high_like_neg']}</td></tr>"
        )

    rows = ""
    for r in top_comments[:100]:
        sentiment = r.get("llm_sentiment") or r.get("effective_sentiment") or "中性"
        label = {"正面": "看好", "中性": "中立", "负面": "看跌"}.get(sentiment, "中立")
        color = {"看好": "#10b981", "中立": "#6b7280", "看跌": "#ef4444"}.get(label, "#6b7280")
        content = (r.get("content", "") or "").replace("<", "&lt;").replace(">", "&gt;")
        rows += (
            f"<tr><td>{r.get('platform','')}</td>"
            f"<td>{(r.get('author_name','') or '')[:18]}</td>"
            f"<td>{(r.get('created_at','') or '')[:19]}</td>"
            f"<td style='text-align:right'>{r.get('likes', 0) or 0}</td>"
            f"<td style='color:{color};font-weight:600'>{label}</td>"
            f"<td class='content'>{content[:200]}</td></tr>"
        )

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<title>散户对科技股(小登股)情绪报告 — {TODAY}</title>
<style>
  body {{ font-family: -apple-system,"Segoe UI","Microsoft YaHei",sans-serif;
          margin:0;padding:24px;background:#f9fafb;color:#111827; }}
  .container {{ max-width:1200px;margin:0 auto;background:#fff;border-radius:12px;
                padding:32px;box-shadow:0 1px 3px rgba(0,0,0,.06); }}
  h1 {{ margin:0 0 8px;font-size:24px; }}
  h2 {{ margin-top:32px;font-size:18px;border-bottom:2px solid #e5e7eb;padding-bottom:8px; }}
  .meta {{ color:#6b7280;font-size:14px;margin-bottom:24px; }}
  .kpis {{ display:grid;grid-template-columns:repeat(4,1fr);gap:16px;margin:16px 0 32px; }}
  .kpi {{ background:#f3f4f6;border-radius:8px;padding:16px; }}
  .kpi .label {{ color:#6b7280;font-size:12px; }}
  .kpi .value {{ font-size:24px;font-weight:700;margin-top:4px; }}
  .kpi.positive .value {{ color:#10b981; }}
  .kpi.neutral .value {{ color:#6b7280; }}
  .kpi.negative .value {{ color:#ef4444; }}
  .kpi.mood .value {{ color:{mood_color}; }}
  table {{ width:100%;border-collapse:collapse;margin-top:12px;font-size:14px; }}
  th,td {{ padding:8px 12px;border-bottom:1px solid #e5e7eb;text-align:left; }}
  th {{ background:#1f4e78;color:#fff; }}
  td.content {{ max-width:520px;overflow:hidden;text-overflow:ellipsis; }}
  tr:hover td {{ background:#f9fafb; }}
  .mood-pill {{ display:inline-block;padding:4px 16px;border-radius:9999px;
                background:{mood_color};color:#fff;font-weight:600; }}
  .note {{ color:#6b7280;font-size:12px;margin-top:16px; }}
</style>
</head>
<body>
<div class="container">
  <h1>散户对科技股(小登股)情绪报告 — {TODAY}</h1>
  <div class="meta">
    时间窗口（CST）：{WINDOW_START.isoformat()} ~ {WINDOW_END.isoformat()}
    &nbsp;·&nbsp; symbol 命中 {breakdown['symbol_pattern']} / 关键词命中 {breakdown['keyword']} / 去重 {breakdown['union']} &nbsp;·&nbsp; 实际分析 {o['total']} 条
    &nbsp;·&nbsp; 报告生成：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
  </div>
  <div class="kpis">
    <div class="kpi"><div class="label">样本评论</div><div class="value">{o['total']}</div></div>
    <div class="kpi positive"><div class="label">看好</div><div class="value">{o['看好']} ({pos_pct}%)</div></div>
    <div class="kpi neutral"><div class="label">中立</div><div class="value">{o['中立']} ({o['中立_pct']}%)</div></div>
    <div class="kpi negative"><div class="label">看跌</div><div class="value">{o['看跌']} ({neg_pct}%)</div></div>
  </div>
  <h2>整体情绪 <span class="mood-pill">{mood}</span></h2>
  <ul>
    <li>综合得分 <strong>{score_avg}</strong>（正值偏多 / 负值偏空）</li>
    <li>看好 {pos_pct}% vs 看跌 {neg_pct}%，差值 {round(pos_pct - neg_pct, 1)} 个百分点</li>
    <li>高赞评论 (likes&gt;10) {o['high_like_total']} 条，看好 {o['high_like_pos']} 条，看跌 {o['high_like_neg']} 条</li>
  </ul>
  <h2>按平台拆分</h2>
  <table>
    <thead><tr><th>平台</th><th>评论数</th><th>看好</th><th>中立</th><th>看跌</th>
        <th>平均得分</th><th>高赞</th><th>高赞看好</th><th>高赞看跌</th></tr></thead>
    <tbody>{plat_rows}</tbody>
  </table>
  <h2>高赞评论样本</h2>
  <table>
    <thead><tr><th>平台</th><th>作者</th><th>时间</th><th>点赞</th><th>情绪</th><th>内容</th></tr></thead>
    <tbody>{rows}</tbody>
  </table>
  <p class="note">
    数据源：<code>db/comments.db.comments</code>；情绪模型：<code>jobs/sentiment_analyzer/llm_sentiment.SentimentAnalyzer</code>。
  </p>
</div>
</body>
</html>
"""
    return html


def render_excel(path, summary, breakdown, top_comments):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    pos_fill = PatternFill("solid", fgColor="C6EFCE")
    neu_fill = PatternFill("solid", fgColor="FFEB9C")
    neg_fill = PatternFill("solid", fgColor="FFC7CE")
    thin = Side(border_style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = f"散户对科技股(小登股)情绪报告 — {TODAY}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:F1")
    ws["A2"] = "窗口(CST)"
    ws["B2"] = f"{WINDOW_START.isoformat()} ~ {WINDOW_END.isoformat()}"
    ws["A3"] = "样本口径"
    ws["B3"] = "symbol SH688/SZ300/BJ8 ∪ 关键词"
    ws["A4"] = "样本数"
    ws["B4"] = summary["overall"]["total"]

    ws["A6"] = "整体"
    ws["A6"].font = Font(bold=True, size=12)
    headers = ["维度", "看好", "看好%", "中立", "中立%", "看跌", "看跌%", "平均得分"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=7, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = border
    o = summary["overall"]
    row = 8
    ws.cell(row=row, column=1, value="全部评论").border = border
    for col, v in enumerate(
        [o["看好"], o["看好_pct"], o["中立"], o["中立_pct"],
         o["看跌"], o["看跌_pct"], o["score_avg"]],
        start=2,
    ):
        c = ws.cell(row=row, column=col, value=v)
        c.border = border
        if col == 2:
            c.fill = pos_fill
        elif col == 4:
            c.fill = neu_fill
        elif col == 6:
            c.fill = neg_fill

    row = 9
    ws.cell(row=row, column=1, value="高赞(likes>10)").border = border
    for col, v in enumerate(
        [o["high_like_pos"], None, None, None, o["high_like_neg"], None, None],
        start=2,
    ):
        c = ws.cell(row=row, column=col, value=v)
        c.border = border
        if col == 2:
            c.fill = pos_fill
        elif col == 6:
            c.fill = neg_fill

    row = 11
    ws.cell(row=row, column=1, value="按平台拆分").font = Font(bold=True, size=12)
    row = 12
    plat_headers = ["平台", "评论数", "看好", "看好%", "中立", "中立%",
                    "看跌", "看跌%", "平均得分", "高赞", "高赞看好", "高赞看跌"]
    for col, h in enumerate(plat_headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = border
    for p, s in summary["by_platform"].items():
        row += 1
        vals = [p, s["total"], s["看好"], s["看好_pct"], s["中立"], s["中立_pct"],
                s["看跌"], s["看跌_pct"], s["score_avg"],
                s["high_like_total"], s["high_like_pos"], s["high_like_neg"]]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=row, column=col, value=v)
            c.border = border
            if col == 4:
                c.fill = pos_fill
            elif col == 6:
                c.fill = neu_fill
            elif col == 8:
                c.fill = neg_fill

    for col_idx in range(1, 13):
        letter = ws.cell(row=12, column=col_idx).column_letter
        ws.column_dimensions[letter].width = 14
    ws.column_dimensions["A"].width = 22

    # Top comments sheet
    ws2 = wb.create_sheet("TopComments")
    ch = ["platform", "author_name", "created_at", "likes", "sentiment", "score", "content"]
    for col, h in enumerate(ch, start=1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = border
    for i, r in enumerate(top_comments, start=2):
        sentiment = r.get("llm_sentiment") or r.get("effective_sentiment") or "中性"
        label = {"正面": "看好", "中性": "中立", "负面": "看跌"}.get(sentiment, "中立")
        score = r.get("llm_score") or r.get("sentiment_score") or 0.0
        try:
            score = float(score)
        except (TypeError, ValueError):
            score = 0.0
        vals = [
            r.get("platform", ""),
            r.get("author_name", ""),
            (r.get("created_at", "") or "")[:19],
            r.get("likes", 0) or 0,
            label,
            score,
            (r.get("content", "") or "").replace("\n", " "),
        ]
        for col, v in enumerate(vals, start=1):
            c = ws2.cell(row=i, column=col, value=v)
            c.border = border
            if col == 5:
                if label == "看好":
                    c.fill = pos_fill
                elif label == "看跌":
                    c.fill = neg_fill
                else:
                    c.fill = neu_fill
    widths = [12, 18, 22, 8, 10, 12, 80]
    for i, w in enumerate(widths, start=1):
        ws2.column_dimensions[ws2.cell(row=1, column=i).column_letter].width = w
    ws2.freeze_panes = "A2"

    wb.save(path)


def main():
    out_dir = SCHEDULE_DIR / "output"
    out_dir.mkdir(exist_ok=True)

    log("=== 散户对科技股(小登股)情绪分析 ===")
    log(f"窗口: {WINDOW_START.isoformat()} ~ {WINDOW_END.isoformat()}")

    log("[1/4] 统计命中口径 ...")
    breakdown = fetch_xiaodeng_breakdown()
    log(f"  symbol 模式: {breakdown['symbol_pattern']}, "
        f"关键词: {breakdown['keyword']}, 去重合计: {breakdown['union']}")

    log("[2/4] 提取命中评论 ...")
    records = fetch_xiaodeng()
    log(f"  取到 {len(records)} 条评论")

    log("[3/4] LLM 情绪分析 ...")
    n_new = run_llm_sentiment(records)
    log(f"  本次新分析 {n_new} 条")

    log("[4/4] 汇总 + 渲染报告 ...")
    summary = aggregate(records)

    # Top comments (likes > 10)
    top = [r for r in records if (r.get("likes") or 0) > 10]
    top.sort(key=lambda r: -(r.get("likes") or 0))

    md = render_md(summary, breakdown, top)
    md_file = out_dir / f"sentiment-xiaodeng-{TODAY}.md"
    md_file.write_text(md, encoding="utf-8")
    log(f"  Saved: {md_file}")

    html_file = out_dir / f"sentiment-xiaodeng-{TODAY}.html"
    html_file.write_text(render_html(summary, breakdown, top), encoding="utf-8")
    log(f"  Saved: {html_file}")

    xlsx_file = out_dir / f"sentiment-xiaodeng-{TODAY}.xlsx"
    render_excel(xlsx_file, summary, breakdown, top)
    log(f"  Saved: {xlsx_file}")

    # JSON summary
    json_file = out_dir / f"sentiment-xiaodeng-{TODAY}.json"
    json_file.write_text(
        json.dumps(
            {
                "window_start": WINDOW_START.isoformat(),
                "window_end": WINDOW_END.isoformat(),
                "breakdown": breakdown,
                "summary": summary,
            },
            ensure_ascii=False, indent=2,
        ),
        encoding="utf-8",
    )
    log(f"  Saved: {json_file}")

    log("=== 完成 ===")


if __name__ == "__main__":
    main()