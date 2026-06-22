#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""按天统计近 30 天散户对老登股(主板蓝筹/红利)的看好 / 中立 / 看跌。

数据源：db/comments.db（窗口内已有 sentiment / sentiment_score，无需新跑 LLM）。
口径：与 `laodeng_sentiment.py` 一致 ——
  - symbol ∈ {SH600*, SH601*, SH603*}（主板大票）
  - symbol IN (laodeng.md 13 个预设标的)
  - content 命中 KEYWORDS 集合（红利/蓝筹/白酒/银行/保险/煤炭/电力/中字头词）
  - 排除：硬科技词 + 互联网/潮玩/茶饮 + 新能源/光伏/锂电

输出：`schedule/collect_all/output/sentiment-laodeng-daily-<TODAY>.{md,html,xlsx,json}`

CLI：
    python laodeng_daily.py                                   # 默认近 30 天，截至 2026-06-20
    python laodeng_daily.py --today 2026-06-23                # 改日期标签
    python laodeng_daily.py --daily-start 2026-06-01 \\
                           --today 2026-06-23                 # 改起始 + 日期标签
"""
import argparse
import io
import json
import sys
from datetime import datetime, timezone, timedelta
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parents[2]            # .../laodeng/ → sentiment_analyzer/ → jobs/ → REPO_ROOT/
SCHEDULE_DIR = REPO_ROOT / "schedule" / "collect_all"
sys.path.insert(0, str(REPO_ROOT))

from backend.database import get_db  # noqa: E402

CST = timezone(timedelta(hours=8))
# 默认：近 30 天（对齐 xiaodeng_daily 6 月窗口起点）
DEFAULT_TODAY = "2026-06-20"
DEFAULT_DAILY_START = "2026-05-23"
TODAY = DEFAULT_TODAY
DAILY_START = DEFAULT_DAILY_START

# 与 laodeng_sentiment.py 完全一致
LAODENG_PRESETS = [
    "SH510050", "SH510300", "SH510500", "SH518800",
    "SH601398", "SH601939", "SH601288", "SH600036",
    "SH600519", "SZ000858", "SZ000568", "SH600900", "SH601088",
]

KEYWORDS = [
    # 大盘 / 宽基 / 板块结构词
    "上证50", "沪深300", "中证500", "中证1000", "蓝筹", "白马", "权重", "大盘",
    "红利", "高股息", "股息率", "央企", "中字头", "国央企",
    # 行业板块词
    "白酒", "消费", "食品饮料", "银行", "保险", "证券", "金融",
    "煤炭", "电力", "公用事业", "石化", "油气", "地产",
    # 行业内具体股票 / 简称
    "茅台", "五粮液", "泸州老窖", "汾酒",
    "工行", "建行", "中行", "招行", "交行",
    "平安", "人寿", "太保", "新华保险",
    "神华", "长江电力", "中石油", "中石化", "中海油",
    # 风格词
    "价值投资", "股息", "长线", "防御",
]


def log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}", flush=True)


def build_where_clause():
    """主板大票 LIKE ∪ 预设 IN ∪ 关键词 OR。"""
    like_clause = (
        "(symbol LIKE 'SH600%' OR symbol LIKE 'SH601%' OR symbol LIKE 'SH603%')"
    )
    placeholders = ",".join(["?"] * len(LAODENG_PRESETS))
    in_clause = f"symbol IN ({placeholders})"
    keyword_clause = " OR ".join(["content LIKE ?"] * len(KEYWORDS))
    sym_params = list(LAODENG_PRESETS)
    kw_params = [f"%{k}%" for k in KEYWORDS]
    return f"({like_clause} OR {in_clause} OR ({keyword_clause}))", sym_params + kw_params


def fetch_daily():
    """返回按天聚合的列表，按日期升序。

    关键指标：
      - score_avg：当日评论数加权平均得分（SentimentAnalyzer 返回的 positive-negative）
      - score_lw：当日 likes 加权平均得分（被赞越多越被广泛认同的方向）
      - delta_score_lw：相对前一日的赞加权得分变化（DoD）
      - cum_score_lw：自 DAILY_START 起的赞加权得分累计（since-DAILY_START）
    """
    where_clause, where_params = build_where_clause()
    sql = f"""
        SELECT date(datetime(created_at, '+8 hours')) AS d,
               COUNT(DISTINCT id) AS n,
               SUM(CASE WHEN COALESCE(sentiment_fix, sentiment) = '正面' THEN 1 ELSE 0 END) AS pos,
               SUM(CASE WHEN COALESCE(sentiment_fix, sentiment) = '中性' THEN 1 ELSE 0 END) AS neu,
               SUM(CASE WHEN COALESCE(sentiment_fix, sentiment) = '负面' THEN 1 ELSE 0 END) AS neg,
               AVG(sentiment_score) AS avg_score,
               SUM(sentiment_score * COALESCE(likes, 0)) AS score_x_likes_sum,
               SUM(COALESCE(likes, 0)) AS likes_sum,
               SUM(CASE WHEN likes > 10 THEN 1 ELSE 0 END) AS hl_total,
               SUM(CASE WHEN likes > 10 AND COALESCE(sentiment_fix, sentiment) = '正面' THEN 1 ELSE 0 END) AS hl_pos,
               SUM(CASE WHEN likes > 10 AND COALESCE(sentiment_fix, sentiment) = '负面' THEN 1 ELSE 0 END) AS hl_neg,
               SUM(CASE WHEN COALESCE(likes, 0) > 0 THEN COALESCE(likes, 0) ELSE 0 END) AS likes_total
        FROM comments
        WHERE created_at IS NOT NULL
          AND date(datetime(created_at, '+8 hours')) >= ?
          AND {where_clause}
        GROUP BY d
        ORDER BY d
    """
    conn = get_db()
    rows = conn.execute(sql, [DAILY_START] + where_params).fetchall()
    conn.close()

    raw = []
    for r in rows:
        d, n, pos, neu, neg, avg_score, score_x_likes_sum, likes_sum, \
            hl_total, hl_pos, hl_neg, likes_total = r
        try:
            avg_score = round(float(avg_score or 0.0), 3)
        except (TypeError, ValueError):
            avg_score = 0.0
        try:
            sxs = float(score_x_likes_sum or 0.0)
            ls = float(likes_sum or 0.0)
            score_lw = round(sxs / ls, 3) if ls > 0 else 0.0
        except (TypeError, ValueError):
            score_lw = 0.0
        raw.append({
            "date": d,
            "total": n or 0,
            "看好": pos or 0,
            "中立": neu or 0,
            "看跌": neg or 0,
            "看好_pct": round(100 * (pos or 0) / n, 1) if n else 0,
            "中立_pct": round(100 * (neu or 0) / n, 1) if n else 0,
            "看跌_pct": round(100 * (neg or 0) / n, 1) if n else 0,
            "score_avg": avg_score,
            "score_lw": score_lw,
            "likes_total": likes_total or 0,
            "high_like_total": hl_total or 0,
            "high_like_pos": hl_pos or 0,
            "high_like_neg": hl_neg or 0,
        })

    # 计算 DoD + 累计赞加权得分（精确版：用 detail 表）
    detail = _fetch_daily_likes_detail()
    detail_map = {d["date"]: d for d in detail}
    cum_sxs = 0.0
    cum_ls = 0.0
    prev_score_lw = None
    for r in raw:
        d = r["date"]
        d_detail = detail_map.get(d, {"sxs": 0.0, "ls": 0.0, "score_lw": 0.0})
        if prev_score_lw is None:
            r["delta_score_lw"] = None
        else:
            r["delta_score_lw"] = round(d_detail["score_lw"] - prev_score_lw, 3)
        prev_score_lw = d_detail["score_lw"]
        cum_sxs += d_detail["sxs"]
        cum_ls += d_detail["ls"]
        r["cum_score_lw"] = round(cum_sxs / cum_ls, 3) if cum_ls > 0 else 0.0
        r["score_lw"] = d_detail["score_lw"]
    return raw


def _fetch_daily_likes_detail():
    """精确聚合：每日 Σ(score × likes) 与 Σ(likes)，用于赞加权得分 + 累计。"""
    where_clause, where_params = build_where_clause()
    sql = f"""
        SELECT date(datetime(created_at, '+8 hours')) AS d,
               SUM(sentiment_score * COALESCE(likes, 0)) AS sxs,
               SUM(COALESCE(likes, 0)) AS ls
        FROM comments
        WHERE created_at IS NOT NULL
          AND date(datetime(created_at, '+8 hours')) >= ?
          AND {where_clause}
        GROUP BY d
        ORDER BY d
    """
    conn = get_db()
    rows = conn.execute(sql, [DAILY_START] + where_params).fetchall()
    conn.close()
    out = []
    for r in rows:
        d, sxs, ls = r
        try:
            sxs = float(sxs or 0.0)
            ls = float(ls or 0.0)
        except (TypeError, ValueError):
            sxs, ls = 0.0, 0.0
        score_lw = round(sxs / ls, 3) if ls > 0 else 0.0
        out.append({"date": d, "sxs": sxs, "ls": ls, "score_lw": score_lw})
    return out


def fetch_platform_daily():
    """按平台 + 天聚合，给拆平台对比用。"""
    where_clause, where_params = build_where_clause()
    sql = f"""
        SELECT date(datetime(created_at, '+8 hours')) AS d,
               platform,
               COUNT(DISTINCT id) AS n,
               SUM(CASE WHEN COALESCE(sentiment_fix, sentiment) = '正面' THEN 1 ELSE 0 END) AS pos,
               SUM(CASE WHEN COALESCE(sentiment_fix, sentiment) = '中性' THEN 1 ELSE 0 END) AS neu,
               SUM(CASE WHEN COALESCE(sentiment_fix, sentiment) = '负面' THEN 1 ELSE 0 END) AS neg,
               AVG(sentiment_score) AS avg_score
        FROM comments
        WHERE created_at IS NOT NULL
          AND date(datetime(created_at, '+8 hours')) >= ?
          AND {where_clause}
        GROUP BY d, platform
        ORDER BY d, platform
    """
    conn = get_db()
    rows = conn.execute(sql, [DAILY_START] + where_params).fetchall()
    conn.close()
    return [
        {
            "date": r[0], "platform": r[1], "total": r[2] or 0,
            "看好": r[3] or 0, "中立": r[4] or 0, "看跌": r[5] or 0,
            "score_avg": round(float(r[6] or 0.0), 3),
        }
        for r in rows
    ]


def _fetch_weekly_likes_detail():
    """每周赞加权得分：每周 Σ(s×likes) / Σ(likes)。"""
    where_clause, where_params = build_where_clause()
    sql = f"""
        SELECT date(datetime(created_at, '+8 hours')) AS d,
               SUM(sentiment_score * COALESCE(likes, 0)) AS sxs,
               SUM(COALESCE(likes, 0)) AS ls
        FROM comments
        WHERE created_at IS NOT NULL
          AND date(datetime(created_at, '+8 hours')) >= ?
          AND {where_clause}
        GROUP BY d
        ORDER BY d
    """
    conn = get_db()
    rows = conn.execute(sql, [DAILY_START] + where_params).fetchall()
    conn.close()

    weekly = {}
    for r in rows:
        d, sxs, ls = r
        try:
            sxs = float(sxs or 0.0)
            ls = float(ls or 0.0)
        except (TypeError, ValueError):
            sxs, ls = 0.0, 0.0
        if not d:
            continue
        dt = datetime.strptime(d, "%Y-%m-%d")
        iso_year, iso_week, _ = dt.isocalendar()
        key = f"{iso_year}-W{iso_week:02d}"
        if key not in weekly:
            weekly[key] = {"sxs": 0.0, "ls": 0.0}
        weekly[key]["sxs"] += sxs
        weekly[key]["ls"] += ls

    out = []
    for key in sorted(weekly.keys()):
        sxs = weekly[key]["sxs"]
        ls = weekly[key]["ls"]
        score_lw = round(sxs / ls, 3) if ls > 0 else 0.0
        out.append({"week": key, "sxs": sxs, "ls": ls, "score_lw": score_lw})
    return out


def _load_xiaodeng_daily_by_date():
    """读小登日报 JSON，按 date → daily row 映射，用于对照表。"""
    xd_json = SCHEDULE_DIR / "output" / f"sentiment-xiaodeng-daily-{TODAY}.json"
    if not xd_json.exists():
        return {}
    try:
        data = json.loads(xd_json.read_text(encoding="utf-8"))
        return {r["date"]: r for r in data.get("daily", []) if r.get("date")}
    except (json.JSONDecodeError, KeyError, OSError) as e:
        log(f"  加载小登日报失败: {e}")
        return {}


def render_md(daily_rows, plat_rows, xiaodeng_by_date):
    total_n = sum(r["total"] for r in daily_rows)
    total_pos = sum(r["看好"] for r in daily_rows)
    total_neu = sum(r["中立"] for r in daily_rows)
    total_neg = sum(r["看跌"] for r in daily_rows)
    if total_n:
        weighted_score = sum(r["score_avg"] * r["total"] for r in daily_rows) / total_n
    else:
        weighted_score = 0.0
    weighted_score = round(weighted_score, 3)
    cum_score_lw = daily_rows[-1]["cum_score_lw"] if daily_rows else 0.0

    lines = [
        f"# 老登股(主板蓝筹/红利)情绪按天统计 — 近 30 天 (截至 {TODAY})",
        "",
        f"- 起始日期：**{DAILY_START}**",
        f"- 样本口径：symbol ∈ `SH600*` / `SH601*` / `SH603*`（主板大票）∪ "
        f"`IN(laodeng.md 13 标的)` ∪ content 命中 {len(KEYWORDS)} 个红利/蓝筹/白酒/银行/保险/煤炭/电力/中字头关键词",
        f"- 累计评论：**{total_n}** 条（看好 {total_pos} / 中立 {total_neu} / 看跌 {total_neg}）",
        f"- 评论数加权得分：**{weighted_score}**（正值偏多 / 负值偏空）",
        f"- **赞加权累计得分（since {DAILY_START}）：{cum_score_lw}**（更看重被广泛认同的方向）",
        f"- 报告生成：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        "## 按天汇总（含赞加权 + 每日变化）",
        "",
        "| 日期 | 评论数 | 看好 | 中立 | 看跌 | 平均得分 | **赞加权得分** | **DoD 变化** | **累计赞加权** | 高赞 |",
        "|------|--------|------|------|------|----------|----------------|--------------|------------------|------|",
    ]
    for r in daily_rows:
        delta = r.get("delta_score_lw")
        delta_str = f"{delta:+g}" if delta is not None else "—"
        lines.append(
            f"| {r['date']} | {r['total']} | {r['看好']} | {r['中立']} | {r['看跌']} | "
            f"{r['score_avg']} | **{r['score_lw']}** | **{delta_str}** | "
            f"**{r['cum_score_lw']}** | {r['high_like_total']} |"
        )
    lines.append("")

    # 趋势观察
    lines.append("## 趋势观察")
    lines.append("")
    if daily_rows:
        total_hl = sum(r["high_like_total"] for r in daily_rows)
        total_hl_pos = sum(r["high_like_pos"] for r in daily_rows)
        total_hl_neg = sum(r["high_like_neg"] for r in daily_rows)
        if total_hl:
            hl_pos_pct = round(100 * total_hl_pos / total_hl, 1)
            lines.append(
                f"- 高赞评论（likes>10）累计 {total_hl} 条，看好 {total_hl_pos} 条（{hl_pos_pct}%），"
                f"看跌 {total_hl_neg} 条（{round(100 * total_hl_neg / total_hl, 1)}%）"
            )

        best_lw = max(daily_rows, key=lambda r: r["score_lw"])
        worst_lw = min(daily_rows, key=lambda r: r["score_lw"])
        lines.append(
            f"- **赞加权得分最高**：{best_lw['date']} = **{best_lw['score_lw']}** "
            f"（{best_lw['total']} 条评论，总赞 {best_lw['likes_total']}）"
        )
        lines.append(
            f"- **赞加权得分最低**：{worst_lw['date']} = **{worst_lw['score_lw']}** "
            f"（{worst_lw['total']} 条评论，总赞 {worst_lw['likes_total']}）"
        )

        deltas = [(r["date"], r["delta_score_lw"]) for r in daily_rows if r.get("delta_score_lw") is not None]
        if deltas:
            up_day = max(deltas, key=lambda x: x[1])
            down_day = min(deltas, key=lambda x: x[1])
            lines.append(
                f"- **当日反转最大**：{up_day[0]}（DoD {up_day[1]:+g}，情绪最强反弹）"
            )
            lines.append(
                f"- **当日回落最大**：{down_day[0]}（DoD {down_day[1]:+g}，情绪最强回调）"
            )

        first_cum = daily_rows[0]["cum_score_lw"]
        last_cum = daily_rows[-1]["cum_score_lw"]
        lines.append(
            f"- **累计赞加权得分变化（since {DAILY_START}）**：{first_cum} → {last_cum} "
            f"（Δ {round(last_cum - first_cum, 3):+g}）"
        )

        # 按 ISO 周汇总
        from collections import OrderedDict
        weekly = OrderedDict()
        for r in daily_rows:
            dt = datetime.strptime(r["date"], "%Y-%m-%d")
            iso_year, iso_week, _ = dt.isocalendar()
            key = f"{iso_year}-W{iso_week:02d}"
            weekly.setdefault(key, {"total": 0, "看好": 0, "中立": 0, "看跌": 0, "score_sum": 0.0})
            weekly[key]["total"] += r["total"]
            weekly[key]["看好"] += r["看好"]
            weekly[key]["中立"] += r["中立"]
            weekly[key]["看跌"] += r["看跌"]
            weekly[key]["score_sum"] += r["score_avg"] * r["total"]

        if weekly:
            lines.append("")
            lines.append("## 按 ISO 周汇总（含赞加权得分）")
            lines.append("")
            lines.append(
                "| 周 | 评论数 | 看好 | 中立 | 看跌 | 看好% | 看跌% | 评论数加权得分 | **赞加权得分** |"
            )
            lines.append(
                "|----|--------|------|------|------|-------|-------|----------------|----------------|"
            )
            weekly_detail = _fetch_weekly_likes_detail()
            weekly_lw_map = {w["week"]: w for w in weekly_detail}
            for wk, w in weekly.items():
                n = w["total"]
                pos_pct = round(100 * w["看好"] / n, 1) if n else 0
                neg_pct = round(100 * w["看跌"] / n, 1) if n else 0
                wavg = round(w["score_sum"] / n, 3) if n else 0.0
                lw = weekly_lw_map.get(wk, {}).get("score_lw", 0.0)
                lines.append(
                    f"| {wk} | {n} | {w['看好']} | {w['中立']} | {w['看跌']} | "
                    f"{pos_pct}% | {neg_pct}% | {wavg} | **{lw}** |"
                )
            lines.append("")

    # 小登 vs 老登 当日得分对照
    if xiaodeng_by_date and daily_rows:
        lines.append("## 小登 vs 老登 当日赞加权得分对照")
        lines.append("")
        lines.append("| 日期 | 小登(硬科技) | 老登(蓝筹/红利) | 差值(老登-小登) | 解读 |")
        lines.append("|------|--------------|-----------------|------------------|------|")
        for r in daily_rows:
            d = r["date"]
            xd_r = xiaodeng_by_date.get(d)
            xd_lw = xd_r["score_lw"] if xd_r else None
            ld_lw = r["score_lw"]
            if xd_lw is None:
                lines.append(
                    f"| {d} | — | {ld_lw} | — | 仅老登有数据 |"
                )
            else:
                diff = round(ld_lw - xd_lw, 3)
                if abs(diff) < 0.05:
                    tag = "同向"
                elif diff > 0.05:
                    tag = "老登更强偏多"
                else:
                    tag = "小登更强偏多"
                lines.append(
                    f"| {d} | {xd_lw} | {ld_lw} | {diff:+g} | {tag} |"
                )
        lines.append("")

    # 按平台 × 天（紧凑）
    if plat_rows:
        from collections import defaultdict
        plat = defaultdict(list)
        for r in plat_rows:
            plat[r["platform"]].append(r)
        lines.append("## 按平台 × 天 (平均得分)")
        lines.append("")
        lines.append("| 日期 | " + " | ".join(plat.keys()) + " |")
        lines.append("|------|" + "|".join(["---"] * len(plat)) + "|")
        all_dates = sorted({r["date"] for r in plat_rows})
        for d in all_dates:
            row = [d]
            for p in plat.keys():
                cell = next((r for r in plat[p] if r["date"] == d), None)
                row.append(f"{cell['score_avg']} ({cell['total']})" if cell else "—")
            lines.append("| " + " | ".join(row) + " |")
        lines.append("")

    lines.append("## 数据源")
    lines.append("")
    lines.append("- 评论表：`db/comments.db.comments`（按 CST 日期 `+8 hours` 偏移聚合）")
    lines.append("- 情绪数据：复用 `db/update_sentiment.py` 已写入的 `sentiment` / `sentiment_fix`")
    lines.append(f"- 关键词集合：{len(KEYWORDS)} 个红利/蓝筹/白酒/银行/保险/煤炭/电力/中字头词 + 风格词（与 `laodeng_sentiment.py` 同步）")
    lines.append("- 预设标的：`data/sections/laodeng.md`（上证 50/300/500/黄金 ETF + 工建中招交 + 茅台五粮液 + 长江电力 + 神华）")
    lines.append("")
    lines.append("## 加权口径说明")
    lines.append("")
    lines.append("- **评论数加权得分**（旧）：`AVG(sentiment_score)` — 当日每条评论得分等权平均")
    lines.append("- **赞加权得分**（新）：`Σ(sentiment_score × likes) / Σ(likes)` — 被点赞越多的评论权重越大，反映'被广泛认同'的方向")
    lines.append("- **DoD 变化**：当日赞加权得分 − 前一日赞加权得分；首日显示 '—'")
    lines.append(f"- **累计赞加权得分**：自 {DAILY_START} 起 Σ(s×likes) / Σ(likes)，单调反映整体情绪累计漂移")
    return "\n".join(lines)


def render_html(daily_rows, plat_rows, xiaodeng_by_date):
    total_n = sum(r["total"] for r in daily_rows)
    cum_score_lw = daily_rows[-1]["cum_score_lw"] if daily_rows else 0.0
    rows = ""
    for r in daily_rows:
        delta = r.get("delta_score_lw")
        delta_str = f"{delta:+g}" if delta is not None else "—"
        delta_color = (
            "#10b981" if (delta is not None and delta > 0)
            else ("#ef4444" if (delta is not None and delta < 0) else "#6b7280")
        )
        lw_color = (
            "#10b981" if r["score_lw"] > 0.05
            else ("#ef4444" if r["score_lw"] < -0.05 else "#6b7280")
        )
        rows += (
            f"<tr><td>{r['date']}</td>"
            f"<td>{r['total']}</td>"
            f"<td style='color:#10b981'>{r['看好']} ({r['看好_pct']}%)</td>"
            f"<td>{r['中立']} ({r['中立_pct']}%)</td>"
            f"<td style='color:#ef4444'>{r['看跌']} ({r['看跌_pct']}%)</td>"
            f"<td>{r['score_avg']}</td>"
            f"<td style='color:{lw_color};font-weight:600'>{r['score_lw']}</td>"
            f"<td style='color:{delta_color};font-weight:600'>{delta_str}</td>"
            f"<td>{r['cum_score_lw']}</td>"
            f"<td>{r['high_like_total']}</td></tr>"
        )

    plat_html = ""
    if plat_rows:
        from collections import defaultdict
        plat = defaultdict(list)
        for r in plat_rows:
            plat[r["platform"]].append(r)
        all_dates = sorted({r["date"] for r in plat_rows})
        plat_html_rows = ""
        for d in all_dates:
            cells = [f"<td>{d}</td>"]
            for p in plat.keys():
                cell = next((r for r in plat[p] if r["date"] == d), None)
                if cell:
                    s = cell["score_avg"]
                    color = "#10b981" if s > 0.05 else ("#ef4444" if s < -0.05 else "#6b7280")
                    cells.append(f"<td style='color:{color}'>{s} <small>({cell['total']})</small></td>")
                else:
                    cells.append("<td>—</td>")
            plat_html_rows += "<tr>" + "".join(cells) + "</tr>"
        plat_html = (
            f"<h2>按平台 × 天 (平均得分)</h2>"
            f"<table><thead><tr><th>日期</th>"
            + "".join(f"<th>{p}</th>" for p in plat.keys())
            + "</tr></thead><tbody>"
            + plat_html_rows
            + "</tbody></table>"
        )

    # 小登 vs 老登 对照表 HTML
    xd_html = ""
    if xiaodeng_by_date and daily_rows:
        xd_rows = ""
        for r in daily_rows:
            d = r["date"]
            xd_r = xiaodeng_by_date.get(d)
            xd_lw = xd_r["score_lw"] if xd_r else None
            ld_lw = r["score_lw"]
            if xd_lw is None:
                xd_rows += f"<tr><td>{d}</td><td>—</td><td>{ld_lw}</td><td>—</td><td>仅老登有数据</td></tr>"
            else:
                diff = round(ld_lw - xd_lw, 3)
                if abs(diff) < 0.05:
                    tag = "同向"
                elif diff > 0.05:
                    tag = "老登更强偏多"
                else:
                    tag = "小登更强偏多"
                diff_color = "#10b981" if diff > 0 else ("#ef4444" if diff < 0 else "#6b7280")
                xd_rows += (
                    f"<tr><td>{d}</td><td>{xd_lw}</td><td>{ld_lw}</td>"
                    f"<td style='color:{diff_color};font-weight:600'>{diff:+g}</td>"
                    f"<td>{tag}</td></tr>"
                )
        xd_html = (
            f"<h2>小登 vs 老登 当日赞加权得分对照</h2>"
            f"<table><thead><tr><th>日期</th><th>小登(硬科技)</th><th>老登(蓝筹/红利)</th>"
            f"<th>差值(老登-小登)</th><th>解读</th></tr></thead>"
            f"<tbody>{xd_rows}</tbody></table>"
        )

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<title>老登股(主板蓝筹/红利)情绪按天统计 — 截至 {TODAY}</title>
<style>
  body {{ font-family: -apple-system,"Segoe UI","Microsoft YaHei",sans-serif;
          margin:0;padding:24px;background:#f9fafb;color:#111827; }}
  .container {{ max-width:1400px;margin:0 auto;background:#fff;border-radius:12px;
                padding:32px;box-shadow:0 1px 3px rgba(0,0,0,.06); }}
  h1 {{ margin:0 0 8px;font-size:24px; }}
  h2 {{ margin-top:32px;font-size:18px;border-bottom:2px solid #e5e7eb;padding-bottom:8px; }}
  .meta {{ color:#6b7280;font-size:14px;margin-bottom:24px; }}
  .kpis {{ display:grid;grid-template-columns:repeat(4,1fr);gap:16px;margin:16px 0 32px; }}
  .kpi {{ background:#f3f4f6;border-radius:8px;padding:16px; }}
  .kpi .label {{ color:#6b7280;font-size:12px; }}
  .kpi .value {{ font-size:24px;font-weight:700;margin-top:4px; }}
  table {{ width:100%;border-collapse:collapse;margin-top:12px;font-size:13px; }}
  th,td {{ padding:8px 12px;border-bottom:1px solid #e5e7eb;text-align:left; }}
  th {{ background:#1f4e78;color:#fff; }}
  tr:hover td {{ background:#f9fafb; }}
  .note {{ color:#6b7280;font-size:12px;margin-top:16px; }}
</style>
</head>
<body>
<div class="container">
  <h1>老登股(主板蓝筹/红利)情绪按天统计 — 截至 {TODAY}</h1>
  <div class="meta">
    起始日期：<strong>{DAILY_START}</strong>
    &nbsp;·&nbsp; 累计评论：<strong>{total_n}</strong> 条
    &nbsp;·&nbsp; <strong>累计赞加权得分：{cum_score_lw}</strong>
    &nbsp;·&nbsp; 报告生成：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
  </div>
  <h2>按天汇总（含赞加权 + 每日变化）</h2>
  <table>
    <thead><tr>
      <th>日期</th><th>评论数</th><th>看好</th><th>中立</th><th>看跌</th>
      <th>平均得分</th><th>赞加权得分</th><th>DoD 变化</th><th>累计赞加权</th><th>高赞</th>
    </tr></thead>
    <tbody>{rows}</tbody>
  </table>
  {xd_html}
  {plat_html}
  <p class="note">
    数据源：<code>db/comments.db.comments</code>；按 CST 日期 <code>+8 hours</code> 偏移聚合；
    情绪复用 <code>db/update_sentiment.py</code> 已写入字段。
    <br>赞加权口径：<code>Σ(score × likes) / Σ(likes)</code>，被点赞越多的评论权重越大，反映"被广泛认同"的方向。
  </p>
</div>
</body>
</html>
"""
    return html


def render_excel(path, daily_rows, plat_rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    pos_fill = PatternFill("solid", fgColor="C6EFCE")
    neu_fill = PatternFill("solid", fgColor="FFEB9C")
    neg_fill = PatternFill("solid", fgColor="FFC7CE")
    pos2_fill = PatternFill("solid", fgColor="D4EDDA")
    neg2_fill = PatternFill("solid", fgColor="F8D7DA")
    thin = Side(border_style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = f"老登股(主板蓝筹/红利)情绪按天统计 — 截至 {TODAY}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:M1")

    headers = [
        "日期", "评论数", "看好", "中立", "看跌",
        "平均得分", "赞加权得分", "DoD 变化", "累计赞加权",
        "高赞", "高赞看好", "高赞看跌",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = border

    for i, r in enumerate(daily_rows, start=4):
        delta = r.get("delta_score_lw")
        vals = [
            r["date"], r["total"], r["看好"], r["中立"], r["看跌"],
            r["score_avg"], r["score_lw"],
            delta if delta is not None else "—",
            r["cum_score_lw"],
            r["high_like_total"], r["high_like_pos"], r["high_like_neg"],
        ]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=i, column=col, value=v)
            c.border = border
            if col == 3:
                c.fill = pos_fill
            elif col == 4:
                c.fill = neu_fill
            elif col == 5:
                c.fill = neg_fill
            elif col == 7:
                if r["score_lw"] > 0.05:
                    c.fill = pos2_fill
                elif r["score_lw"] < -0.05:
                    c.fill = neg2_fill
            elif col == 8:
                if isinstance(delta, (int, float)):
                    if delta > 0.05:
                        c.fill = pos2_fill
                    elif delta < -0.05:
                        c.fill = neg2_fill

    # Per-platform sheet
    if plat_rows:
        ws2 = wb.create_sheet("Platform×Day")
        ph = ["日期", "平台", "评论数", "看好", "中立", "看跌", "平均得分"]
        for col, h in enumerate(ph, start=1):
            c = ws2.cell(row=1, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center")
            c.border = border
        for i, r in enumerate(plat_rows, start=2):
            vals = [r["date"], r["platform"], r["total"], r["看好"], r["中立"],
                    r["看跌"], r["score_avg"]]
            for col, v in enumerate(vals, start=1):
                c = ws2.cell(row=i, column=col, value=v)
                c.border = border
                if col == 4:
                    c.fill = pos_fill
                elif col == 5:
                    c.fill = neu_fill
                elif col == 6:
                    c.fill = neg_fill

    # 列宽
    for col_idx in range(1, 14):
        letter = ws.cell(row=3, column=col_idx).column_letter
        ws.column_dimensions[letter].width = 12
    ws.column_dimensions["A"].width = 14

    wb.save(path)


def parse_args(argv=None):
    p = argparse.ArgumentParser(description="老登股(主板蓝筹/红利)情绪按天统计")
    p.add_argument("--daily-start", help="起始日期 YYYY-MM-DD（默认 2026-05-23）",
                   default=None)
    p.add_argument("--today", help="日期标签 YYYY-MM-DD（默认 2026-06-20）", default=None)
    return p.parse_args(argv)


def apply_args(args):
    """把 CLI 参数覆盖到模块级 DAILY_START / TODAY。"""
    global DAILY_START, TODAY
    DAILY_START = args.daily_start or DEFAULT_DAILY_START
    TODAY = args.today or DEFAULT_TODAY


def main():
    apply_args(parse_args())

    out_dir = SCHEDULE_DIR / "output"
    out_dir.mkdir(exist_ok=True)

    log("=== 老登股(主板蓝筹/红利)情绪按天统计 ===")
    log(f"起始日期: {DAILY_START}")

    log("[1/4] 拉取按天聚合 ...")
    daily = fetch_daily()
    log(f"  共 {len(daily)} 天, 累计 {sum(r['total'] for r in daily)} 条")

    log("[2/4] 拉取按平台 × 天聚合 ...")
    plat = fetch_platform_daily()
    log(f"  共 {len(plat)} 行")

    log("[3/4] 加载小登日报（用于对照） ...")
    xiaodeng_by_date = _load_xiaodeng_daily_by_date()
    log(f"  小登日报共 {len(xiaodeng_by_date)} 天数据")

    log("[4/4] 渲染报告 ...")
    md = render_md(daily, plat, xiaodeng_by_date)
    md_file = out_dir / f"sentiment-laodeng-daily-{TODAY}.md"
    md_file.write_text(md, encoding="utf-8")
    log(f"  Saved: {md_file}")

    html_file = out_dir / f"sentiment-laodeng-daily-{TODAY}.html"
    html_file.write_text(render_html(daily, plat, xiaodeng_by_date), encoding="utf-8")
    log(f"  Saved: {html_file}")

    xlsx_file = out_dir / f"sentiment-laodeng-daily-{TODAY}.xlsx"
    render_excel(xlsx_file, daily, plat)
    log(f"  Saved: {xlsx_file}")

    json_file = out_dir / f"sentiment-laodeng-daily-{TODAY}.json"
    json_file.write_text(
        json.dumps(
            {
                "start_date": DAILY_START,
                "today": TODAY,
                "daily": daily,
                "platform_daily": plat,
            },
            ensure_ascii=False, indent=2,
        ),
        encoding="utf-8",
    )
    log(f"  Saved: {json_file}")

    log("=== 完成 ===")


if __name__ == "__main__":
    main()